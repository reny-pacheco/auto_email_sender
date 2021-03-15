import openpyxl

def get_data(excel_filename):
    students_sheet = openpyxl.load_workbook(excel_filename)
    students_data = students_sheet['Sheet1']

    students = []
    email = []

    for student_row in range(2, students_data.max_row + 1):
        student_name = students_data.cell(student_row, 2).value.split(" ")[0]
        student_email = students_data.cell(student_row, 3).value

        students.append(student_name)
        email.append(student_email)

    return students, email